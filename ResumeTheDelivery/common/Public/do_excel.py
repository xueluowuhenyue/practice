from openpyxl import load_workbook
from common.Public import project_path
from common.Public.read_ini import Configuration


class DoExcel:
    def __init__(self, file_name):
        self.file_name=file_name

    def write_back(self, sheet_name, row, cloumn, value):
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, cloumn, value)
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    p = DoExcel(project_path.Excel_path)
    p.write_back('info', 2, 1, 333)

