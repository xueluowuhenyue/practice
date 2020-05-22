# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from API_C.Common.public.read_ini import configuration
from  API_C.Common.public import project_path
# from openpyxl import Workbook
button=configuration(project_path.config_path).read_str('CASE','button')
class DoExcel:
    '''这个类的作用是完成Excel数据的读写 新建表单的操作'''
    def __init__(self,file_name):
        self.file_name=file_name
    def read_Excel(self,sheet_name):
        '''读取指定表单的数据'''
        wb=load_workbook(self.file_name)
        sheet=wb[sheet_name]
        '''新建列表用于接收所有数据'''
        max_list=[]
        '''遍历Excel文件中表单的数据'''
        for i in range(2,sheet.max_row+1):
            row_list=[]
            '''新建列表用于接收每行的数据'''
            for j in range(1,sheet.max_column+1):
                '''将数据添加到列表中'''
                row_list.append(sheet.cell(i,j).value)
            max_list.append(row_list)
            #执行配置文件中的测试用例
            list=[]
            if button=='all':
                list=max_list
            else:
                for i in max_list:
                    if i[0] in eval(button):            #读取单独测试用例
                    # if i[1] in eval(button):            #按照功能模块读取用例
                        list.append(i)
            if row_list[5].find('number') != -1:
                '''获取表单中的电话号码'''
                phone=DoExcel(project_path.excel_path).read_phone()
                '''替换test表单中的电话号码'''
                row_list[5]=row_list[5].replace('number',str(phone))
                self.undate_phone(phone+1)
            else:
                row_list[5]=row_list[5]
        return list
        wb.close()
    def read_phone(self,sheet_name='number'):
        '''用来查看电话号码'''
        wb=load_workbook(self.file_name)
        '''定位表单'''
        sheet=wb[sheet_name]
        ph=sheet.cell(2,3).value
        return ph
        wb.close()
    def undate_phone(self,value,sheet_name='number'):
        '''用来更新话号码'''
        wb=load_workbook(self.file_name)
        '''定位表单'''
        sheet=wb[sheet_name]
        '''写入数据'''
        sheet.cell(2,3,value)
        wb.save(self.file_name)
        wb.close()
    def write_Excel(self,sheet_name,row,column,value):
        '''在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb=load_workbook(self.file_name)
        '''定位表单'''
        sheet=wb[sheet_name]
        # logger.info('开始写入数据啦')
        '''写入数据'''
        sheet.cell(row,column,value)
        '''保存'''
        wb.save(self.file_name)
        '''关闭文件'''
        wb.close()

if __name__ == '__main__':
    p=DoExcel(project_path.excel_path)
    # print(p.read_phone())
    print(p.read_Excel('test'))
