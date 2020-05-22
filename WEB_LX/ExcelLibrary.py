# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class ExcelLibrary:
    # ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    # ROBOT_LIBRARY_VERSION = 1.0
    # def __init__(self,file_name, sheet_name):
    #     self.file_name = file_name + '.xlsx'
    #     self.sheet_name = sheet_name
    def Create_Excel_File(self,file_name):
        """
        创建Excel文件
        :param file_name:
        :return:
        """
        wb = Workbook()
        wb.save(file_name+'.xlsx')

    def Load_Excel_File(self, file_name, sheet_name):
        """打开文件/定位表单
        :param file_name:
        :param sheet_name:
        :return:
        """
        WB = load_workbook(file_name)
        sheet = WB[sheet_name]
        return sheet

    def Get_Max_Row(self, file_name, sheet_name):
        """获取表单最大行数
        :param file_name:
        :param sheet_name:
        :return:
        """
        sheet = self.Load_Excel_File(file_name, sheet_name)
        return sheet.max_row

    def Get_Max_Column(self, file_name, sheet_name):
        """获取表单最大列数
        :param file_name:
        :param sheet_name:
        :return:
        """
        sheet = self.Load_Excel_File(file_name, sheet_name)
        return sheet.max_column

    def Read_Row_Data(self, file_name, sheet_name, row):
        """读取指定行数据
        :param file_name:
        :param sheet_name:
        :param row:
        :return:
        """
        col=self.Get_Max_Column(file_name, sheet_name)
        list=[]
        for i in range(1, col+1):
            list.append(self.Load_Excel_File(file_name, sheet_name).cell(int(row), i).value)
        return list

    def Read_Column_Data(self, file_name, sheet_name, column):
        """
        读取指定行数据
        :param file_name:
        :param sheet_name:
        :param column:
        :return:
        """
        row=self.Get_Max_Row(file_name, sheet_name)
        list=[]
        for i in range(2, row+1):
            list.append(self.Load_Excel_File(file_name, sheet_name).cell(i, int(column)).value)
        return list

    def Read_Cell_Data(self, file_name, sheet_name, row, column):
        """读取指定单元格数据
        :param file_name:
        :param sheet_name:
        :param row:
        :param column:
        :return:
        """
        data=self.Load_Excel_File(file_name, sheet_name).cell(int(row), int(column)).value
        return data

    def Read_Excel_Data(self,file_name, sheet_name):
        """读取表格所有数据
        """
        sheet = self.Load_Excel_File(file_name, sheet_name)
        row=self.Get_Max_Row(file_name, sheet_name)
        col=self.Get_Max_Column(file_name, sheet_name)
        max_list = []
        for i in range(1, row+1):
            row_list = []
            for j in range(1, col+1):
                row_list.append(sheet.cell(i, j).value)
            max_list.append(row_list)
        return max_list

    def Set_Cell_Data(self, file_name, sheet_name,  row, column, value):
        """
        修改/添加表单数据
        :param file_name:
        :param sheet_name:
        :param row:
        :param column:
        :param value:
        :return:
        """
        wb = load_workbook(file_name)
        '''定位表单'''
        sheet = wb[sheet_name]
        '''写入数据'''
        sheet.cell(int(row), int(column), value)
        wb.save(file_name)
        wb.close()


if __name__ == '__main__':
    mm=ExcelLibrary()
    print(mm.Get_Max_Row('log.xlsx','register'))
    print(mm.Get_Max_Column('log.xlsx','register'))
    print(mm.Read_Row_Data('log.xlsx','register',3))
    print(mm.Read_Column_Data('log.xlsx','register',1))
    print(mm.Read_Cell_Data('log.xlsx','register',1,2))
    print(mm.Read_Excel_Data('log.xlsx','register'))
    print(mm.Set_Cell_Data('log.xlsx','register',1,2,666))
    mm.Create_Excel_File('kn')