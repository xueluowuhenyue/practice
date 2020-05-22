# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from Common.Public import Project_path

class DoExcel:
    def __init__(self, file_name):
        self.file_name = file_name

    def read_data(self, sheet_name):
        ' ''数据读取'' '
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        data = []
        for i in range(2, sheet.max_row+1):
            row_list = []
            for j in range(1, sheet.max_column+1):
                row_list.append(sheet.cell(i, j).value)
            data.append(row_list)
            wb.close()
        return data

    def write_data(self, sheet_name, row, column, value):
        ' ''数据写入'' '
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, column, value)
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    res=DoExcel(Project_path.Excel_path).read_data('login')
    print(res)