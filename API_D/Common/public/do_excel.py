from openpyxl import load_workbook
from API_D.Common.public import project_path
from API_D.Common.public.read_ini import configuration
button=configuration(project_path.config_path).read_str('CASE','button')


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_data(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        max_data=[]
        for i in range(2, sheet.max_row+1):
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['URL'] = sheet.cell(i, 3).value
            row_data['Title'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            row_data['Params'] = sheet.cell(i, 6).value
            row_data['Sql'] = sheet.cell(i, 7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            max_data.append(row_data)
            # 执行配置文件中的测试用例
            old_list = []
            if button == 'all':
                old_list = max_data
            else:
                for i in max_data:
                    # if i['CaseId'] in eval(button):            # 读取单独测试用例
                    if i['Module'] in eval(button):  # 按照功能模块读取用例
                        old_list.append(i)
            if row_data['Params'].find('number') != -1:
                '''获取表单中的电话号码'''
                phone = DoExcel(project_path.excel_path).read_phone()
                '''替换test表单中的电话号码'''
                row_data['Params'] = row_data['Params'].replace('number', str(phone))
                self.undate_phone(phone + 1)
            else:
                row_data['Params'] = row_data['Params']
        wb.close()
        return old_list

    def read_phone(self, sheet_name='number'):
        '''用来查看电话号码'''
        wb = load_workbook(self.file_name)
        '''定位表单'''
        sheet = wb[sheet_name]
        ph = sheet.cell(2, 3).value
        wb.close()
        return ph

    def undate_phone(self, value, sheet_name='number'):
        '''用来更新话号码'''
        wb = load_workbook(self.file_name)
        '''定位表单'''
        sheet = wb[sheet_name]
        '''写入数据'''
        sheet.cell(2, 3, value)
        wb.save(self.file_name)
        wb.close()

    def write_back(self,row,cloumn,value):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,cloumn,value)
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    p = DoExcel(project_path.excel_path,'test')
    res=p.read_data()
    print(res)
