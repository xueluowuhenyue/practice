# -*- coding:utf-8 -*-
#openpyxl  同时支持Excel的读写  xlrd  xlwt
#只支持：xlsx   Excel

from openpyxl import load_workbook  #读写
from openpyxl import Workbook       #新建

#新建Excel文件
# wb=Workbook()     #创建
# wb.save('py14.xlsx')   #保存，否则创建失败

#打开文件
wb=load_workbook('py14.xlsx')         #路径

#定位表单
# sheet=wb.get_sheet_by_name('test')  #根据名字定位表单 会警告
sheet=wb['test']                      #现在用

#定位单元格
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        res=sheet.cell(row=i,column=j).value
        if res == None:
            pass
        else:
            print(res)

#定位表单 写值
#一
sheet.cell(row=5,column=5,value=True)
wb.save('py14.xlsx')

res=sheet.cell(row=2,column=4).value
print(type(res))
# 二
# sheet.cell(row=6,column=6).value='失败'
# wb.save('py14.xlsx')

#获取最大行和列
# print('行',sheet.max_row)
# print('列',sheet.max_column)

