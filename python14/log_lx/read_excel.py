# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from python14.log_lx.conf_read import configuration
from python14.log_lx.logging_class import Mylog

logger=Mylog()
class DoExcel:
    '''这个类的作用是完成Excel数据的读写 新建表单的操作'''
    def __init__(self,file_name,sheet_name,row=None,column=None):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.row=row
        self.column=column
    def read_Excel(self):
        '''读取指定表单的数据'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        '''新建列表用于接收所有数据'''
        list=[]
        '''遍历Excel文件中表单的数据'''
        logger.info('开始读取数据啦')
        for i in range(1,sheet.max_row+1):
            row_list=[]
            '''新建列表用于接收每行的数据'''
            for j in range(1,sheet.max_column+1):
                '''将数据添加到列表中'''
                row_list.append(sheet.cell(i,j).value)
            list.append(row_list)
        print(list)
        logger.info('读取数据读取完毕啦')

    def write_Excel(self,value):
        '''在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb=load_workbook(self.file_name)
        '''定位表单'''
        sheet=wb[self.sheet_name]
        logger.info('开始写入数据啦')
        try:
            '''写入数据'''
            sheet.cell(self.row,self.column,value)
            '''保存'''
            wb.save(self.file_name)
            '''关闭文件'''
            wb.close()
        except Exception as e:
            logger.error(e)
        finally:
            logger.info('数据写入完毕')
    def creat_Excel(self):
        '''新建一个Excel'''
        logger.info('开始创建Excel表')
        wb=Workbook()
        try:
            '''新建表单'''
            wb.create_sheet(self.sheet_name)
            '''新建Excel文件'''
            wb.save(self.file_name)
        except Exception as e:
            logger.error(e)
            raise e
        finally:
            logger.info('表格创建成功')

if __name__ == '__main__':
    p=DoExcel('D:\pycharm\python14\log_lx\py14.xlsx','test')   #不传行、列参数
    # p.creat_Excel()
    p.read_Excel()
    # p.write_Excel('kk')