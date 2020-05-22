# -*- coding: utf-8 -*-
# @Time    : 2019/2/22 20:11
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : learn_excel.py

#python 的openpyxl模块来操作 Excel
#为什么要操作Excel
#禅道  xmind  excel

#pip install openpyxl
#openpyxl 里面 load_workbook 是打开Excel文件的
#workbook sheet cell
#学习的时候 可以类比open函数
#操作的Excel xlsx结尾
from openpyxl import load_workbook
# wb=load_workbook('D:\\2018Python自动化课件&代码\code\python_14\week_6\class_0222\python_14.xlsx')
#打开工作簿
# wb=load_workbook('python_14.xlsx')
#
# #定位表单 wb[表单名]
# sheet=wb['Sheet1']
# # sheet_1=wb.get_sheet_by_name('Sheet1')
#
# #定位单元格 excel的行列值是从1开始数的
# # res=sheet.cell(column=1,row=4).value
#
# #可以得到最大的行数 列数
# # print(sheet.max_row)
# # print(sheet.max_column)
#
# #请你把第六行的数据全部读出来
# # for i in range(1,sheet.max_column+1):
# #     res=sheet.cell(column=i,row=6).value
# #     print(res)
#
# #所有行所有列的数据全部都读出去？
# for i in range(1,sheet.max_row+1):#i=1 i=2 i=3 i=4 i=5 i=6
#     for j in range(1,sheet.max_column+1):#j=1 2 3
#         res=sheet.cell(row=i,column=j).value
#         if res!=None:
#             print(res)
# #读取Excel的时候 返回的是None  那么就说明你有空的单元格
#
# #写数据：
# # sheet.cell(3,2,'作者：李白')#行 列 value
# sheet.cell(3,2).value='作者：柠檬班小高'
# #只要进行了单元格的更改都要去进行保存
# wb.save('python_14.xlsx')
#
#
# #新建工作簿
from openpyxl import workbook

wb=workbook.Workbook()
wb.save('python16.xlsx')
#循环写数据 写到不同的表单里面去  类与对象的问题（写成类与对象再去拓展）

# def add(a,b):
#     print(a+b)
#
# add(1,2)
# add(a=1,b=2)
# add(b=2,a=1)
