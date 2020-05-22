# -*- coding: utf-8 -*-
# @Time    : 2019/2/22 21:36
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py
#写一个类 类的作用是完成Excel数据的读写 新建表单的操作
#函数一：读取指定表单的数据，
#有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。
#每一行都有 一个单独的row_list [[1,2,3],[4,5,6]]
#每一行数据读取完毕后，把row_list存到大列表all_row_list
#函数二：在指定的单元格写入指定的数据，并保存到当前Excel
#函数三：新建一个Excel
from openpyxl import  workbook
from openpyxl import  load_workbook

class DoExcel:
    '''类的作用是完成Excel数据的读写 新建表单的操作'''

    def fun_1(self):
        '''读取所有的数据，以嵌套列表的形式存储，每一行都是一个子列表，每一个单元格都是子列表里面的元素'''
        pass#留给你们的作业

    def write_excel(self,file_name,sheet_name,row,col,value):
        '''在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]
        sheet.cell(row,col).value=value
        wb.save(file_name)

    def create_excel(self,file_name,sheet_name):
        '''新建一个Excel'''
        wb=workbook.Workbook()
        wb.create_sheet(sheet_name)#新建表单
        wb.save(file_name)

if __name__ == '__main__':
    DoExcel().create_excel('python20.xlsx','lemon')
    DoExcel().write_excel('python_14.xlsx','Sheet1',5,2,'信Python得永生')