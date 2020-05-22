# -*- coding: utf-8 -*-
# @Time    : 2019/2/22 21:36
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : do_excel.py
#写一个类 类的作用是完成Excel数据的读写 新建表单的操作
#函数一：读取指定表单的数据，
#有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。
#每一行都有 一个单独的row_list [[1,2,3],[4,5,6]]
#每一行数据读取完毕后，把row_list存到大列表all_row_list
#函数二：在指定的单元格写入指定的数据，并保存到当前Excel
#函数三：新建一个Excel

from openpyxl import  workbook
from openpyxl import  load_workbook
from week_7.class_0227.read_config import ReadConfig#用这个模块  要用我们刚刚写的类
from week_7.class_0227.my_log import MyLog

logger=MyLog()

class DoExcel:
    '''类的作用是完成Excel数据的读写 新建表单的操作'''

    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_excel(self,button):#
        '''读取所有的数据，以嵌套列表的形式存储，每一行都是一个子列表，每一个单元格都是子列表里面的元素'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]

        #嵌套列表--嵌套循环
        # test_data=[]#大列表 所有的字列表会存在这个里面
        # for j in range(2,sheet.max_row+1):
        #     row_data=[]#每一行数据存在一个字列表里面
        #     for i in range(1,sheet.max_column+1):
        #         row_data.append(sheet.cell(j,i).value)
        #     test_data.append(row_data)
        # return test_data
        #

        #单层循环
        logger.info('开始读取数据了啦！')
        test_data=[]#大列表 所有的字列表会存在这个里面
        if button=='1':#1读取所有的用例
            for i in range(2,sheet.max_row+1):
                row_data=[]#每一行数据存在一个字列表里面
                row_data.append(sheet.cell(i,1).value)
                row_data.append(sheet.cell(i,2).value)
                row_data.append(sheet.cell(i,3).value)
                row_data.append(sheet.cell(i,4).value)
                row_data.append(sheet.cell(i,5).value)
                row_data.append(sheet.cell(i,6).value)
                row_data.append(sheet.cell(i,7).value)
                test_data.append(row_data)
            logger.info('读取数据完毕！')
        else:
            test_data=[]#大列表 所有的字列表会存在这个里面
            for i in eval(button):#如果button不等于1  eval()之后就是一个列表
                row_data=[]#每一行数据存在一个字列表里面
                row_data.append(sheet.cell(i+1,1).value)
                row_data.append(sheet.cell(i+1,2).value)
                row_data.append(sheet.cell(i+1,3).value)
                row_data.append(sheet.cell(i+1,4).value)
                row_data.append(sheet.cell(i+1,5).value)
                row_data.append(sheet.cell(i+1,6).value)
                row_data.append(sheet.cell(i+1,7).value)
                test_data.append(row_data)
            logger.info('读取数据完毕！')

        return test_data

        #嵌套字典
        # test_data=[]#大列表 所有的字字典会存在这个里面
        # for i in range(2,sheet.max_row+1):
        #     row_data={}#每一行数据存在一个字典里面
        #     row_data['CaseId']=sheet.cell(i,1).value
        #     row_data['Title']=sheet.cell(i,2).value
        #     row_data['Module']=sheet.cell(i,3).value
        #     row_data['TestData']=sheet.cell(i,4).value
        #     row_data['ExpectedResult']=sheet.cell(i,5).value
        #     row_data['ActualResult']=sheet.cell(i,6).value
        #     row_data['TestReuslt']=sheet.cell(i,7).value
        #     test_data.append(row_data)
        # return test_data

    def write_excel(self,row,col,value):
        '''在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        logger.info('开始往Excel里面写数据')
        try:
            sheet.cell(row,col).value=value
            wb.save(self.file_name)
            wb.close()#每次操作完 关闭掉！！！
        except Exception as e:
            logger.error(e)
        logger.info('Excel里面数据写入完毕')

    def create_excel(self):
        '''新建一个Excel'''
        wb=workbook.Workbook()
        wb.create_sheet(self.sheet_name)#新建表单
        wb.save(self.file_name)

if __name__ == '__main__':
    button=ReadConfig('lemon.conf').get_data('TestCase','button')
    test_data=DoExcel('python_14.xlsx','test_case').read_excel(button)
    print(test_data)