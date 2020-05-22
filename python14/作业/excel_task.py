# # -*- coding:utf-8 -*-
# #1：写一个类的作用是完成Excel数据的读写 新建表单的操作
# #函数一：读取指定表单的数据,有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。
# # #每一行都有 一个单独的row_list [[1,2,3],[4,5,6]]
# # #每一行数据读取完毕后，把row_list存到大列表all_row_list
# # #函数二：在指定的单元格写入指定的数据，并保存到当前Excel
# # #函数三：新建一个Excel
from openpyxl import load_workbook
from openpyxl import Workbook

class DoExcel:
    '''这个类的作用是完成Excel数据的读写 新建表单的操作'''
    def __init__(self,file_name,sheet_name,row=None,column=None):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.row=row
        self.column=column
    def read_Excel(self):
        '''读取指定表单的数据'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        '''新建列表用于接收所有数据'''
        list=[]
        '''遍历Excel文件中表单的数据'''
        for i in range(1,sheet.max_row+1):
            row_list=[]
            '''新建列表用于接收每行的数据'''
            for j in range(1,sheet.max_column+1):
                '''将数据添加到列表中'''
                row_list.append(sheet.cell(i,j).value)
            list.append(row_list)
        print(list)
        '''数据查询'''
        print(list[self.row-1][self.column-1])
    def write_Excel(self,value):
        '''在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb=load_workbook(self.file_name)
        '''定位表单'''
        sheet=wb[self.sheet_name]
        '''写入数据'''
        res=sheet.cell(self.row,self.column,value)
        '''保存'''
        wb.save(self.file_name)
    def creat_Excel(self):
        '''新建一个Excel'''
        wb=Workbook()
        '''新建表单'''
        wb.create_sheet(self.sheet_name)
        '''新建Excel文件'''
        wb.save(self.file_name)
if __name__ == '__main__':
    p=DoExcel('D:\pycharm\python14\week5\py14.xlsx','test',2,2)
    # p.creat_Excel()
    p.read_Excel()
    # p.write_Excel('kk')
